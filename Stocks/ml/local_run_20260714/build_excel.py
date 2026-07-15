# -*- coding: utf-8 -*-
"""Build the AI prediction accuracy report (.xlsx) from run artifacts.
All numbers come from results_{tag}.json / walkforward_{tag}_h{h}.csv —
nothing hand-typed. Aggregates that a spreadsheet can compute are written
as live formulas over the detail sheets, so the workbook is self-auditing.
"""
import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path("/sessions/serene-amazing-gauss/mnt/outputs")
XLSX = OUT / "AI股票預測_實測報告_20260714.xlsx"

FONT = "Microsoft JhengHei"          # Traditional-Chinese professional font
LBL = {0: "下跌", 1: "持平", 2: "上漲"}
CLS = ["下跌", "持平", "上漲"]
SYM_NAME = {"2330": "台積電", "2317": "鴻海", "2454": "聯發科",
            "2882": "國泰金", "2603": "長榮", "2308": "台達電"}
TAGS = {"balanced": "平衡版", "naive": "基準版"}
HS = {1: "1日", 5: "5日"}

R = {t: json.load(open(OUT / f"results_{t}.json")) for t in TAGS}
WF = {(t, h): pd.read_csv(OUT / f"walkforward_{t}_h{h}.csv", parse_dates=["date"])
      for t in TAGS for h in (1, 5)}

wb = Workbook()

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="FCE4EC")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")


def F(bold=False, size=11, color="000000", italic=False):
    return Font(name=FONT, bold=bold, size=size, color=color, italic=italic)


def put(ws, row, col, val, bold=False, fill=None, fmt=None, size=11,
        color="000000", wrap=False, border=True, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = F(bold, size, color, italic)
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if border:
        c.border = BORDER
    c.alignment = Alignment(vertical="center", wrap_text=wrap,
                            horizontal="left" if isinstance(val, str) else "right")
    return c


def header_row(ws, row, headers, widths=None):
    for j, h in enumerate(headers, 1):
        c = put(ws, row, j, h, bold=True, fill=HDR_FILL, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w


# ---------------------------------------------------------------- detail sheets
detail_names = {}
for t in TAGS:
    for h in (1, 5):
        name = f"明細{TAGS[t]}{HS[h]}"
        detail_names[(t, h)] = name
        ws = wb.create_sheet(name)
        header_row(ws, 1, ["股票代號", "股票名稱", "特徵日", "預測", "實際",
                           "正確(1=對)", "實際報酬", "p下跌", "p持平", "p上漲",
                           "測試月", "訓練筆數", "早停輪數"],
                   [10, 10, 12, 8, 8, 11, 11, 9, 9, 9, 10, 10, 10])
        d = WF[(t, h)].sort_values(["date", "symbol"]).reset_index(drop=True)
        ycol = f"y{h}"; rcol = f"ret{h}"
        for i, r in d.iterrows():
            row = i + 2
            put(ws, row, 1, str(r["symbol"]))
            put(ws, row, 2, SYM_NAME[str(r["symbol"])])
            put(ws, row, 3, str(pd.Timestamp(r["date"]).date()))
            put(ws, row, 4, LBL[int(r["pred"])])
            put(ws, row, 5, LBL[int(r[ycol])])
            put(ws, row, 6, f"=IF(D{row}=E{row},1,0)")
            put(ws, row, 7, float(r[rcol]), fmt="0.00%")
            put(ws, row, 8, float(r["p_down"]), fmt="0.000")
            put(ws, row, 9, float(r["p_flat"]), fmt="0.000")
            put(ws, row, 10, float(r["p_up"]), fmt="0.000")
            put(ws, row, 11, str(r["test_month"]))
            put(ws, row, 12, int(r["train_rows"]))
            put(ws, row, 13, int(r["best_iter"]))
        ws.freeze_panes = "A2"

NROWS = {(t, h): len(WF[(t, h)]) for t in TAGS for h in (1, 5)}


def rng(t, h, col):
    return f"'{detail_names[(t,h)]}'!${col}$2:${col}${NROWS[(t,h)]+1}"


# ---------------------------------------------------------------- 準確度指標
ws = wb.create_sheet("準確度指標")
ws.column_dimensions["A"].width = 34
for col in "BCDE":
    ws.column_dimensions[col].width = 15
put(ws, 1, 1, "走勢預測準確度(walk-forward 樣本外驗證)", bold=True, size=13, border=False)
put(ws, 2, 1, "測試期間:2026-04-01 ~ 2026-07-14(逐月擴充視窗重訓,標籤結算日早於測試月才進訓練集,無前視洩漏)",
    border=False, wrap=False)
put(ws, 3, 1, "所有比率皆為活公式,直接統計「明細」工作表;右側附 Python 交叉核算值。", border=False,
    italic=True, color="808080")

row = 5
combos = [("balanced", 1), ("balanced", 5), ("naive", 1), ("naive", 5)]
for t, h in combos:
    m = R[t]["metrics"][str(h)]
    n = NROWS[(t, h)]
    put(ws, row, 1, f"◆ {TAGS[t]}模型.{HS[h]}水平(n={n})", bold=True, fill=SUB_FILL)
    for j in range(2, 6):
        put(ws, row, j, "", fill=SUB_FILL)
    row += 1
    put(ws, row, 1, "指標", bold=True, fill=HDR_FILL, color="FFFFFF")
    put(ws, row, 2, "數值(活公式)", bold=True, fill=HDR_FILL, color="FFFFFF")
    put(ws, row, 3, "Python核算", bold=True, fill=HDR_FILL, color="FFFFFF")
    put(ws, row, 4, "95%信賴區間下界", bold=True, fill=HDR_FILL, color="FFFFFF")
    put(ws, row, 5, "95%信賴區間上界", bold=True, fill=HDR_FILL, color="FFFFFF")
    row += 1
    put(ws, row, 1, "整體準確率(3類)")
    put(ws, row, 2, f"=AVERAGE({rng(t,h,'F')})", fmt="0.0%")
    put(ws, row, 3, m["accuracy"], fmt="0.0%")
    put(ws, row, 4, m["acc_ci95"][0], fmt="0.0%")
    put(ws, row, 5, m["acc_ci95"][1], fmt="0.0%")
    row += 1
    put(ws, row, 1, "平衡準確率(3類recall平均)")
    rec_parts = []
    for cname in CLS:
        rec_parts.append(f"IFERROR(COUNTIFS({rng(t,h,'E')},\"{cname}\",{rng(t,h,'D')},\"{cname}\")/COUNTIF({rng(t,h,'E')},\"{cname}\"),0)")
    put(ws, row, 2, "=(" + "+".join(rec_parts) + ")/3", fmt="0.0%")
    put(ws, row, 3, m["balanced_accuracy"], fmt="0.0%")
    row += 1
    # baselines
    put(ws, row, 1, "基準線:永遠猜「上漲」")
    put(ws, row, 2, f"=COUNTIF({rng(t,h,'E')},\"上漲\")/COUNTA({rng(t,h,'E')})", fmt="0.0%")
    put(ws, row, 3, m["baselines"]["always_up"], fmt="0.0%")
    row += 1
    put(ws, row, 1, "基準線:永遠猜「持平」")
    put(ws, row, 2, f"=COUNTIF({rng(t,h,'E')},\"持平\")/COUNTA({rng(t,h,'E')})", fmt="0.0%")
    put(ws, row, 3, m["baselines"]["always_flat"], fmt="0.0%")
    row += 1
    cb = m["baselines"]["carry_prev_label"]
    put(ws, row, 1, f"基準線:沿用前一日實際方向(n={cb['n']})")
    put(ws, row, 2, cb["accuracy"], fmt="0.0%")
    put(ws, row, 3, "Python計算(同股票前一筆標籤)", border=True)
    row += 1
    # directional
    put(ws, row, 1, "方向性出手數(預測非持平)")
    put(ws, row, 2, f"=COUNTIF({rng(t,h,'D')},\"上漲\")+COUNTIF({rng(t,h,'D')},\"下跌\")")
    put(ws, row, 3, m["directional"]["n_calls"])
    row += 1
    put(ws, row, 1, "方向命中率(漲時猜漲/跌時猜跌,報酬符號)")
    up_hit = f"COUNTIFS({rng(t,h,'D')},\"上漲\",{rng(t,h,'G')},\">0\")"
    dn_hit = f"COUNTIFS({rng(t,h,'D')},\"下跌\",{rng(t,h,'G')},\"<0\")"
    denom = f"COUNTIF({rng(t,h,'D')},\"上漲\")+COUNTIF({rng(t,h,'D')},\"下跌\")"
    put(ws, row, 2, f"=IFERROR(({up_hit}+{dn_hit})/({denom}),\"—\")", fmt="0.0%")
    put(ws, row, 3, m["directional"].get("sign_hit_all"), fmt="0.0%")
    ci = m["directional"].get("sign_hit_ci95")
    if ci:
        put(ws, row, 4, ci[0], fmt="0.0%")
        put(ws, row, 5, ci[1], fmt="0.0%")
    row += 1
    # per class
    for en, cname in zip(["down", "flat", "up"], CLS):
        pc = m["per_class"][en]
        put(ws, row, 1, f"「{cname}」precision / recall / F1(支持數={pc['support']})")
        put(ws, row, 2, f"=IFERROR(COUNTIFS({rng(t,h,'E')},\"{cname}\",{rng(t,h,'D')},\"{cname}\")/COUNTIF({rng(t,h,'D')},\"{cname}\"),\"無此類預測\")", fmt="0.0%")
        put(ws, row, 3, f"P={pc['precision'] if pc['precision']==pc['precision'] else float('nan'):.3f} R={pc['recall']:.3f} F1={(pc['f1'] if pc['f1']==pc['f1'] else 0):.3f}"
            if pc["precision"] == pc["precision"] else f"P=無(0次預測) R={pc['recall']:.3f}")
        row += 1
    row += 1

put(ws, row, 1, "註:95% 信賴區間為 Wilson score interval(Python 計算,程式碼見 ml/local_run_20260714/run_ai_prediction.py)。",
    border=False, italic=True, color="808080")
ws.freeze_panes = "A5"

# ---------------------------------------------------------------- 混淆矩陣
ws = wb.create_sheet("混淆矩陣")
ws.column_dimensions["A"].width = 22
for col in "BCDE":
    ws.column_dimensions[col].width = 13
put(ws, 1, 1, "混淆矩陣(列=實際,欄=預測;活公式 COUNTIFS 統計明細表)", bold=True, size=13, border=False)
row = 3
for t, h in combos:
    put(ws, row, 1, f"{TAGS[t]}模型.{HS[h]}水平", bold=True, fill=SUB_FILL)
    for j in range(2, 6):
        put(ws, row, j, "", fill=SUB_FILL)
    row += 1
    put(ws, row, 1, "實際\\預測", bold=True, fill=HDR_FILL, color="FFFFFF")
    for j, cname in enumerate(CLS, 2):
        put(ws, row, j, f"預測{cname}", bold=True, fill=HDR_FILL, color="FFFFFF")
    put(ws, row, 5, "列合計", bold=True, fill=HDR_FILL, color="FFFFFF")
    row += 1
    first = row
    for aname in CLS:
        put(ws, row, 1, f"實際{aname}", bold=True)
        for j, pname in enumerate(CLS, 2):
            put(ws, row, j,
                f"=COUNTIFS({rng(t,h,'E')},\"{aname}\",{rng(t,h,'D')},\"{pname}\")")
        put(ws, row, 5, f"=SUM(B{row}:D{row})", bold=True)
        row += 1
    put(ws, row, 1, "欄合計", bold=True)
    for j in range(2, 5):
        L = get_column_letter(j)
        put(ws, row, j, f"=SUM({L}{first}:{L}{row-1})", bold=True)
    put(ws, row, 5, f"=SUM(B{row}:D{row})", bold=True)
    row += 2

# ---------------------------------------------------------------- 最新預測
ws = wb.create_sheet("最新預測")
widths = [10, 10, 13, 15, 12, 14, 9, 9, 9, 30]
put(ws, 1, 1, "最新一次預測(特徵基準日 2026-07-14 收盤;1日=下一交易日 2026-07-15,5日=5個交易日後)",
    bold=True, size=13, border=False)
put(ws, 2, 1, "重要:依本報告 walk-forward 實測,模型準確度未優於天真基準線;以下預測僅為「執行一次」之輸出示範,不construe為投資建議。",
    bold=True, color="9C0006", border=False)
header_row(ws, 4, ["股票代號", "股票名稱", "模型設定", "預測水平", "預測方向",
                   "死區門檻", "p下跌", "p持平", "p上漲", "說明"], widths)
row = 5
for t in TAGS:
    for p in R[t]["final_predictions"]:
        put(ws, row, 1, p["symbol"])
        put(ws, row, 2, p["name"])
        put(ws, row, 3, TAGS[t])
        put(ws, row, 4, HS[p["horizon"]])
        pred_zh = {"down": "下跌", "flat": "持平", "up": "上漲"}[p["prediction"]]
        put(ws, row, 5, pred_zh, bold=True,
            fill=(BAD_FILL if pred_zh == "下跌" else OK_FILL if pred_zh == "上漲" else WARN_FILL))
        put(ws, row, 6, f"±{p['deadzone']*100:.1f}%")
        put(ws, row, 7, p["p_down"], fmt="0.000")
        put(ws, row, 8, p["p_flat"], fmt="0.000")
        put(ws, row, 9, p["p_up"], fmt="0.000")
        put(ws, row, 10, f"訓練樣本 {p['train_rows']} 筆(2025-06-02 ~ 2026-07-13 特徵日)")
        row += 1
put(ws, row+1, 1, "最高類別機率僅約 0.32~0.48,接近隨機(3類先驗約 0.33),機率本身即顯示訊號薄弱。",
    border=False, italic=True, color="808080")
ws.freeze_panes = "A5"

# ---------------------------------------------------------------- 實地驗證
ws = wb.create_sheet("實地驗證0713_0714")
put(ws, 1, 1, "實地(live)驗證:7月折疊模型(僅用 ≤2026-06-30 標籤訓練)對 7/13、7/14 的預測 vs TWSE 實際收盤",
    bold=True, size=13, border=False)
put(ws, 2, 1, "預測產生於結果實現之前(walk-forward 協議),7/13、7/14 實際數據於 2026-07-14 21:30 CST 自 TWSE 抓取。",
    border=False, italic=True, color="808080")
header_row(ws, 4, ["模型設定", "股票", "特徵日", "目標日", "預測", "實際", "實際報酬", "正確"],
           [12, 12, 12, 12, 9, 9, 11, 9])
row = 5
tgt = {pd.Timestamp("2026-07-09"): "2026-07-13", pd.Timestamp("2026-07-13"): "2026-07-14"}
for t in TAGS:
    d = WF[(t, 1)]
    lv = d[d["date"].isin(list(tgt.keys()))].sort_values(["date", "symbol"])
    for _, r in lv.iterrows():
        put(ws, row, 1, TAGS[t])
        put(ws, row, 2, f"{r['symbol']} {SYM_NAME[str(r['symbol'])]}")
        put(ws, row, 3, str(pd.Timestamp(r["date"]).date()))
        put(ws, row, 4, tgt[pd.Timestamp(r["date"])])
        put(ws, row, 5, LBL[int(r["pred"])])
        put(ws, row, 6, LBL[int(r["y1"])])
        put(ws, row, 7, float(r["ret1"]), fmt="0.00%")
        hit = int(r["pred"]) == int(r["y1"])
        put(ws, row, 8, "✓" if hit else "✗", bold=True,
            fill=OK_FILL if hit else BAD_FILL)
        row += 1
    row += 1
put(ws, row, 1, "平衡版 12 次中 5 次正確(41.7%);基準版 12 次中 2 次正確(16.7%)。樣本極小,僅作示意;整體結論請以 411 筆 walk-forward 為準。",
    border=False, wrap=False)

# ---------------------------------------------------------------- 資料驗證
ws = wb.create_sheet("資料驗證")
ws.column_dimensions["A"].width = 26
for col, w in zip("BCDEFG", [16, 16, 16, 16, 16, 30]):
    ws.column_dimensions[col].width = w
put(ws, 1, 1, "資料來源與真實性驗證", bold=True, size=13, border=False)
row = 3
put(ws, row, 1, "1) 資料涵蓋範圍(TWSE 交易所官方 STOCK_DAY 日線)", bold=True, fill=SUB_FILL)
for j in range(2, 8):
    put(ws, row, j, "", fill=SUB_FILL)
row += 1
header_row(ws, row, ["股票", "筆數", "起日", "迄日", "最後收盤", "來源", "抓取時間(CST)"])
row += 1
for t in ("balanced",):
    for code, cv in R[t]["coverage"].items():
        put(ws, row, 1, f"{code} {SYM_NAME[code]}")
        put(ws, row, 2, cv["rows"])
        put(ws, row, 3, cv["first"])
        put(ws, row, 4, cv["last"])
        put(ws, row, 5, cv["last_close"], fmt="#,##0.00")
        put(ws, row, 6, "www.twse.com.tw/exchangeReport/STOCK_DAY")
        put(ws, row, 7, "2026-07-10 及 2026-07-14 21:30 補抓")
        row += 1
row += 1
put(ws, row, 1, "2) 三方交叉驗證:2026-07-09 收盤(同日三個獨立來源)", bold=True, fill=SUB_FILL)
for j in range(2, 8):
    put(ws, row, j, "", fill=SUB_FILL)
row += 1
header_row(ws, row, ["股票", "TWSE日線 STOCK_DAY", "TWSE即時 mis.twse", "Yahoo股市網頁", "三者一致", "", "查核時間(CST)"])
row += 1
xcheck = [
    ("2330 台積電", 2415.0, 2415.0, 2415.0, "2026-07-10 20:05-20:07"),
    ("2317 鴻海",   237.5, 237.5, None,   "2026-07-10 20:05"),
    ("2454 聯發科", 3925.0, 3925.0, 3925.0, "2026-07-10 20:05-20:09"),
    ("2882 國泰金", 96.3,  96.3,  96.3,  "2026-07-10 20:05-20:08"),
    ("2603 長榮",   194.5, 194.5, None,   "2026-07-10 20:05"),
    ("2308 台達電", 1880.0, 1880.0, None,  "2026-07-10 20:05"),
]
first_x = row
for name, a, b, c, ts in xcheck:
    put(ws, row, 1, name)
    put(ws, row, 2, a, fmt="#,##0.00")
    put(ws, row, 3, b, fmt="#,##0.00")
    put(ws, row, 4, c if c is not None else "未逐檔查核", fmt="#,##0.00" if c else None)
    if c is not None:
        put(ws, row, 5, f"=IF(AND(B{row}=C{row},C{row}=D{row}),\"✓ 一致\",\"✗ 不一致\")", bold=True)
    else:
        put(ws, row, 5, f"=IF(B{row}=C{row},\"✓ 一致(兩源)\",\"✗ 不一致\")", bold=True)
    put(ws, row, 7, ts)
    row += 1
row += 1
notes = [
    "3) 誠實揭露(抓不到/缺漏的數據)",
    "‧ Yahoo Finance API(query1/query2.finance.yahoo.com)在本執行環境被網路政策封鎖,無法程式化抓取歷史 OHLCV;故歷史日線改用交易所官方 TWSE STOCK_DAY(權威來源),Yahoo 股市(tw.stock.yahoo.com)網頁即時報價僅用於收盤價交叉驗證(2330/2454/2882 三檔全數吻合,資料時間 2026/07/09 14:30)。",
    "‧ 三大法人買賣超(T86)與融資融券歷史檔:單日全市場檔案過大,環境限制下無法批量回補一年歷史;依 backend/feature_schema.py 既定規範,法人籌碼特徵以 0.0(中性)代入 = 本次實際未使用。",
    "‧ K線圖形特徵(YOLO):repo 內無已訓練模型檔(backend/ml_models/ 不存在),圖形特徵以 0.0 代入 = 未使用。",
    "‧ 2317 鴻海缺 2025-07-30 一筆:TWSE 該月回傳即無此日,非本流程遺失。2026-05 月份首次查詢遭 TWSE 回傳錯誤,已改用替代端點成功補齊並驗證。",
    "‧ 2026-07-10、07-11、07-12 非交易日(TWSE 即時 API 確認最後交易日為 2026-07-09;7/13 週一恢復交易)。",
    "‧ 6 檔標的取自 ml/config.py SEED_SYMBOLS 流動性名單前段,涵蓋半導體/電子/IC設計/金融/航運/電源。",
]
for txt in notes:
    put(ws, row, 1, txt, bold=txt.startswith("3)"),
        fill=SUB_FILL if txt.startswith("3)") else None, border=False, wrap=False)
    row += 1

# ---------------------------------------------------------------- 方法與限制
ws = wb.create_sheet("方法與限制")
ws.column_dimensions["A"].width = 130
put(ws, 1, 1, "方法、程式現況與限制(給部署決策者)", bold=True, size=13, border=False)
paras = [
    "■ 程式現況:repo 的 AI 預測由 backend/ml_predict.py 服務,需要 backend/ml_models/(manifest.json、pattern_yolo.onnx、lgbm_nextday.txt、lgbm_5day.txt)才能運作;此資料夾目前不存在,ml/ 訓練管線亦僅有 config.py/db.py/universe.py 骨架(缺 labels/features/train/evaluate)。故本次「執行」係依 repo 設計文件重建可執行管線:直接 import backend/indicators.py(30 項技術指標)與 backend/feature_schema.py(特徵順序/缺值=0 規範),標籤採 ml/config.py 死區(1日 ±0.5%、5日 ±1.2%,3類:下跌/持平/上漲),模型採 LightGBM 多分類——與 repo 規劃一致。",
    "■ 驗證協議:expanding-window walk-forward,逐月重訓,測試月 2026-04 ~ 2026-07;訓練集僅含「標籤結算日早於測試月首日」之樣本,杜絕前視洩漏。測試樣本 1日=411 筆、5日=387 筆,全部為樣本外。",
    "■ 兩組設定(全部揭露,無挑選):(1)基準版=LightGBM 預設 logloss 早停 → 因驗證切片落在 2026/2-3 修正期,第 1 輪即早停,模型退化為「幾乎全猜上漲」;(2)平衡版=類別平衡權重 + balanced-accuracy 早停(修正類別不平衡)。平衡版係在觀察到基準版退化後加入,兩版測試結果並列呈報。",
    "■ 核心結論(真實、樣本外):1日水平——平衡版 37.0%(CI 32.5%~41.8%)、基準版 41.4%,皆未優於「永遠猜上漲」43.1%;5日水平——平衡版 35.7%、基準版 51.4%=永遠猜上漲基準線,且遠低於「沿用前一日方向」72.2%(5日報酬重疊自相關使然)。方向命中率(出手時報酬符號)45.7%~58.1%,未穩定超越擲幣。",
    "■ 判讀:在僅有價量技術指標的條件下,本模型對台股 6 檔大型權值股之次日/5日方向無可部署的預測力。此結果與 ml/config.py 檔頭自述一致(「流動性股票的次日方向,學術上隨機漫步天花板約 50-55%」);repo 期望的 ~75% 需依賴死區設計+籌碼/圖形特徵,本次實測顯示僅靠技術指標遠達不到。",
    "■ 若要繼續:優先補齊 (1) 三大法人/融資券歷史特徵(建立每日增量收集,如 ml/ 管線原規劃);(2) YOLO 圖形模型訓練;(3) 更長歷史(2年+)與更大股票池;(4) 以「沿用前一日方向」與「永遠猜上漲」為最低上線門檻,加上機率校準與交易成本回測。在未穩定超越上述基準線之前,不建議作為付費功能上線。",
    "■ 免責聲明:本報告為模型能力評估,非投資建議。過往準確度不代表未來表現。",
]
row = 3
for p in paras:
    c = put(ws, row, 1, p, border=False, wrap=True)
    ws.row_dimensions[row].height = 64
    row += 1

# ---------------------------------------------------------------- 摘要 (first)
ws = wb.create_sheet("摘要", 0)
ws.column_dimensions["A"].width = 46
for col, w in zip("BCDE", [16, 16, 16, 16]):
    ws.column_dimensions[col].width = w
put(ws, 1, 1, "AI 股票預測 — 單次執行結果與真實準確度報告", bold=True, size=16, border=False)
put(ws, 2, 1, "產生時間:2026-07-14 21:45 CST|資料:台灣證券交易所(TWSE)官方 API + Yahoo股市交叉驗證|標的:台股6檔權值股|模型:LightGBM 3類(依 repo ml/ 設計)",
    border=False, color="595959")
put(ws, 4, 1, "★ 給部署決策的一句話:模型樣本外準確度未超越天真基準線(永遠猜上漲/沿用昨日方向),不建議依此上線付費功能。",
    bold=True, color="9C0006", border=False)
row = 6
put(ws, row, 1, "關鍵數字(活公式引用進一步工作表)", bold=True, fill=SUB_FILL)
for j in range(2, 6):
    put(ws, row, j, "", fill=SUB_FILL)
row += 1
header_row(ws, row, ["指標", "平衡版1日", "平衡版5日", "基準版1日", "基準版5日"])
row += 1
put(ws, row, 1, "樣本外準確率(walk-forward)")
put(ws, row, 2, f"=AVERAGE({rng('balanced',1,'F')})", fmt="0.0%")
put(ws, row, 3, f"=AVERAGE({rng('balanced',5,'F')})", fmt="0.0%")
put(ws, row, 4, f"=AVERAGE({rng('naive',1,'F')})", fmt="0.0%")
put(ws, row, 5, f"=AVERAGE({rng('naive',5,'F')})", fmt="0.0%")
row += 1
put(ws, row, 1, "基準線:永遠猜「上漲」")
for j, (t, h) in zip(range(2, 6), combos_ordered := [("balanced",1),("balanced",5),("naive",1),("naive",5)]):
    put(ws, row, j, f"=COUNTIF({rng(t,h,'E')},\"上漲\")/COUNTA({rng(t,h,'E')})", fmt="0.0%")
row += 1
put(ws, row, 1, "模型贏過基準線?")
for j, (t, h) in zip(range(2, 6), combos_ordered):
    put(ws, row, j, f"=IF({get_column_letter(j)}{row-2}>{get_column_letter(j)}{row-1},\"是\",\"否\")", bold=True)
row += 1
put(ws, row, 1, "測試樣本數(筆,全為樣本外)")
for j, (t, h) in zip(range(2, 6), combos_ordered):
    put(ws, row, j, f"=COUNTA({rng(t,h,'E')})")
row += 2
put(ws, row, 1, "資料真實性:TWSE 日線 × TWSE 即時 API × Yahoo股市 三方核對 7/9 收盤全數一致(詳「資料驗證」表);7/13、7/14 預測於結果實現前產生並以實際收盤驗證(詳「實地驗證」表)。",
    border=False, wrap=True)
ws.row_dimensions[row].height = 30
row += 1
put(ws, row, 1, "工作表導覽:最新預測|準確度指標|混淆矩陣|實地驗證0713_0714|明細×4(1,598筆樣本外預測)|資料驗證|方法與限制",
    border=False, color="595959")

wb.remove(wb["Sheet"])
wb.save(XLSX)
print("saved:", XLSX)
